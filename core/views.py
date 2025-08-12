from django.db.models import Q
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.http import JsonResponse, FileResponse, HttpResponse, HttpResponseForbidden, Http404 # Dùng để trả về file Excel
import io # Dùng để làm việc với file trong bộ nhớ
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages # Để hiển thị thông báo cho người dùng
from django.contrib.auth.decorators import login_required # Dành cho function-based views
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from users.models import CustomUser
from django.core.exceptions import PermissionDenied

from django.urls import reverse_lazy
from django.views.decorators.http import require_POST
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.utils.translation import gettext_lazy as _

from .models import (
    CoQuan, PhongBan, HoSoCongViec, KeHoach, MocThoiGian, NhiemVu,
    TepDinhKem, LichSuCongViec, CustomReport, Notification
)
# Assuming forms exist in core/forms.py
from .forms import (
    CoQuanForm, PhongBanForm, HoSoCongViecForm, KeHoachForm, MocThoiGianForm,
    NhiemVuForm, TepDinhKemForm, LichSuCongViecForm, CustomReportForm, TaskHandoverForm, ExtensionRequestForm, BinhLuanForm, NhiemVuCompletionForm
)

from datetime import date, timedelta, datetime
from django.utils import timezone
from .utils import create_notification, calculate_business_hours # Import the utility function
import json # For parsing JSON in update_task_date_from_calendar

@login_required
def get_user_role(request, user_id):
    try:
        user = CustomUser.objects.get(pk=user_id)
        return JsonResponse({'role': user.role})
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)

def home(request):
    return render(request, 'core/home.html')

@login_required
def dashboard(request):
    user = request.user
    context = {'user': user}
    today = date.today()

    # Determine base_tasks_queryset based on user's role (reusing NhiemVuListView's logic)
    if user.is_superuser or user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
        base_tasks_queryset = NhiemVu.objects.all()
        context['all_projects'] = HoSoCongViec.objects.all() # Keep for superuser/LANH_DAO_CO_QUAN
        context['all_plans'] = KeHoach.objects.all() # Keep for superuser/LANH_DAO_CO_QUAN
        context['all_tasks'] = NhiemVu.objects.all() # Keep for superuser/LANH_DAO_CO_QUAN
        if user.co_quan: # Keep for LANH_DAO_CO_QUAN
            context['agency_projects'] = HoSoCongViec.objects.filter(id_don_vi_chu_tri__co_quan=user.co_quan)
            context['agency_plans'] = KeHoach.objects.filter(id_don_vi_thuc_hien__co_quan=user.co_quan)
            context['agency_tasks'] = NhiemVu.objects.filter(id_phong_ban_lien_quan__co_quan=user.co_quan)
        else: # Keep for LANH_DAO_CO_QUAN
            context['agency_projects'] = HoSoCongViec.objects.none()
            context['agency_plans'] = KeHoach.objects.none()
            context['agency_tasks'] = NhiemVu.objects.none()
    elif user.role == CustomUser.Role.LANH_DAO_VAN_PHONG:
        if user.phong_ban and user.phong_ban.co_quan:
            base_tasks_queryset = NhiemVu.objects.filter(
                id_phong_ban_lien_quan__co_quan=user.phong_ban.co_quan
            )
        else:
            base_tasks_queryset = NhiemVu.objects.none()
    elif user.role == CustomUser.Role.LANH_DAO_PHONG:
        if user.phong_ban:
            base_tasks_queryset = NhiemVu.objects.filter(id_phong_ban_lien_quan=user.phong_ban)
            # Keep department-specific data for LANH_DAO_PHONG
            phong_ban = user.phong_ban
            context['department_projects'] = HoSoCongViec.objects.filter(id_don_vi_chu_tri=phong_ban)
            context['department_plans'] = KeHoach.objects.filter(id_du_an__id_don_vi_chu_tri=phong_ban)
            context['department_tasks'] = NhiemVu.objects.filter(id_ke_hoach__id_du_an__id_don_vi_chu_tri=phong_ban)
        else:
            base_tasks_queryset = NhiemVu.objects.none()
            context.update({ # Keep for LANH_DAO_PHONG
                'department_projects': HoSoCongViec.objects.none(),
                'department_plans': KeHoach.objects.none(),
                'department_tasks': NhiemVu.objects.none()
            })
    elif user.role == CustomUser.Role.CHUYEN_VIEN_VAN_PHONG or user.role == CustomUser.Role.CHUYEN_VIEN_PHONG:
        base_tasks_queryset = NhiemVu.objects.filter(Q(id_nguoi_xu_ly_chinh=user) | Q(id_nguoi_giao_viec=user) | Q(nguoi_dong_xu_ly=user)).distinct()
        context['assigned_tasks'] = NhiemVu.objects.filter(Q(id_nguoi_xu_ly_chinh=user) | Q(nguoi_dong_xu_ly=user)).distinct() # Keep for CHUYEN_VIEN
    else:
        base_tasks_queryset = NhiemVu.objects.none()

    # Populate tasks_due_in_one_day and upcoming_tasks for all roles
    tomorrow = today + timedelta(days=1)
    context['tasks_due_in_one_day'] = base_tasks_queryset.filter(
        ngay_ket_thuc__date=tomorrow,
        trang_thai__in=[NhiemVu.TrangThai.PENDING_ASSIGNMENT_APPROVAL, NhiemVu.TrangThai.ASSIGNED, NhiemVu.TrangThai.PENDING_COMPLETION_APPROVAL]
    ).order_by('ngay_ket_thuc')

    context['upcoming_tasks'] = base_tasks_queryset.filter(
        ngay_ket_thuc__gte=today,
        ngay_ket_thuc__date__gt=tomorrow, # Exclude tasks due tomorrow
        trang_thai__in=[NhiemVu.TrangThai.PENDING_ASSIGNMENT_APPROVAL, NhiemVu.TrangThai.ASSIGNED, NhiemVu.TrangThai.PENDING_COMPLETION_APPROVAL]
    ).order_by('ngay_ket_thuc') # Sort by remaining days in ascending order

    # Query for items needing assignment approval
    pending_assignment_tasks = NhiemVu.objects.none()
    pending_assignment_hosocongviec = HoSoCongViec.objects.none()
    pending_assignment_kehoach = KeHoach.objects.none()

    if user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
        pending_assignment_tasks = NhiemVu.objects.filter(trang_thai=NhiemVu.TrangThai.PENDING_ASSIGNMENT_APPROVAL)
        pending_assignment_hosocongviec = HoSoCongViec.objects.filter(trang_thai=HoSoCongViec.TrangThai.CHO_PHE_DUYET)
        # For KeHoach, if LANH_DAO_CO_QUAN can approve initial plans
        pending_assignment_kehoach = KeHoach.objects.filter(trang_thai=KeHoach.TrangThai.CHUA_BAT_DAU)
    elif user.role == CustomUser.Role.LANH_DAO_VAN_PHONG:
        if user.phong_ban and user.phong_ban.co_quan:
            pending_assignment_tasks = NhiemVu.objects.filter(
                trang_thai=NhiemVu.TrangThai.PENDING_ASSIGNMENT_APPROVAL,
                id_phong_ban_lien_quan__co_quan=user.phong_ban.co_quan
            )
            pending_assignment_hosocongviec = HoSoCongViec.objects.filter(
                trang_thai=HoSoCongViec.TrangThai.CHO_PHE_DUYET,
                id_don_vi_chu_tri__co_quan=user.phong_ban.co_quan
            )
            # For KeHoach, if LANH_DAO_VAN_PHONG can approve initial plans within their agency
            pending_assignment_kehoach = KeHoach.objects.filter(
                trang_thai=KeHoach.TrangThai.CHUA_BAT_DAU,
                id_don_vi_thuc_hien__co_quan=user.phong_ban.co_quan
            )
    elif user.role == CustomUser.Role.LANH_DAO_PHONG:
        if user.phong_ban:
            # Lãnh đạo Phòng approves NhiemVu within their department
            pending_assignment_tasks = NhiemVu.objects.filter(
                trang_thai=NhiemVu.TrangThai.PENDING_ASSIGNMENT_APPROVAL,
                id_phong_ban_lien_quan=user.phong_ban
            )
            # Lãnh đạo Phòng approves KeHoach within their department
            pending_assignment_kehoach = KeHoach.objects.filter(
                trang_thai=KeHoach.TrangThai.CHUA_BAT_DAU,
                id_don_vi_thuc_hien=user.phong_ban
            )
            # HoSoCongViec is approved by LANH_DAO_CO_QUAN, so LANH_DAO_PHONG won't see it here for assignment approval.
            # If they need to see it, the logic for HoSoCongViec approval needs to be adjusted.

    # For completion approval, the approver is specific to the task (NhiemVu only for now)
    pending_completion_tasks = NhiemVu.objects.filter(
        trang_thai=NhiemVu.TrangThai.PENDING_COMPLETION_APPROVAL
    ).filter(
        Q(is_quy_trinh_dac_biet=True, id_nguoi_duyet_giao_viec=user) |
        Q(is_quy_trinh_dac_biet=False, id_nguoi_giao_viec=user)
    )

    context['pending_assignment_tasks'] = pending_assignment_tasks
    context['pending_assignment_hosocongviec'] = pending_assignment_hosocongviec
    context['pending_assignment_kehoach'] = pending_assignment_kehoach
    context['pending_completion_tasks'] = pending_completion_tasks

    return render(request, 'core/dashboard.html', context)

@login_required
def nhiemvu_kanban_view(request):
    context = {}
    return render(request, 'core/nhiemvu_kanban.html', context)

@login_required
def nhiemvu_calendar_view(request):
    context = {}
    return render(request, 'core/nhiemvu_calendar.html', context)

@login_required
def nhiemvu_calendar_data(request):
    events = []
    return JsonResponse(events, safe=False)

@login_required
def nhiemvu_gantt_view(request):
    context = {}
    return render(request, 'core/nhiemvu_gantt.html', context)

@login_required
def nhiemvu_gantt_data(request):
    data = []
    return JsonResponse({'data': data})

@login_required
def report_project_progress_view(request):
    return render(request, 'core/report_project_progress.html', {})

@login_required
def export_project_progress_pdf(request):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer)
    p.drawString(inch, 10.5 * inch, "Báo cáo tiến độ dự án")
    p.line(inch, 10.4 * inch, 7.5 * inch, 10.4 * inch)
    p.showPage()
    p.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='bao_cao_tien_do.pdf')

@login_required
def export_project_progress_excel(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Báo cáo tiến độ'
    headers = ['Tên dự án', 'Tiến độ (%)', 'Tổng số nhiệm vụ', 'Nhiệm vụ hoàn thành']
    for col_num, header_title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header_title
        cell.font = Font(bold=True)
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].autosize = True
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=bao_cao_tien_do.xlsx'
    workbook.save(response)
    return response

@login_required
def notification_list(request):
    notifications = Notification.objects.filter(user=request.user)
    context = {
        'notifications': notifications
    }
    return render(request, 'core/notification_list.html', context)

@login_required
def mark_notification_as_read(request, pk):
    notification = get_object_or_404(Notification, pk=pk, user=request.user)
    if not notification.is_read:
        notification.is_read = True
        notification.save()
    return redirect('notification-list')

@login_required
def get_unread_notifications_count(request):
    count = Notification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({'unread_count': count})

@login_required
def get_users_for_autocomplete(request):
    search_term = request.GET.get('q', '')
    users = CustomUser.objects.filter(
        Q(username__icontains=search_term) | Q(first_name__icontains=search_term) | Q(last_name__icontains=search_term)
    ).exclude(role=CustomUser.Role.LANH_DAO_CO_QUAN).order_by('username')[:10] # Limit to 10 results

    results = []
    for user in users:
        results.append({'id': user.pk, 'text': user.get_full_name() or user.username})
    return JsonResponse({'results': results})

@require_POST
@login_required
def update_task_date_from_calendar(request):
    try:
        data = json.loads(request.body)
        task_id = data.get('id')
        new_start_str = data.get('start')
        new_end_str = data.get('end')
        if not all([task_id, new_start_str, new_end_str]):
            return JsonResponse({'status': 'error', 'message': 'Dữ liệu không đủ.'}, status=400)
        task = get_object_or_404(NhiemVu, pk=task_id)
        # Add permission check here if necessary
        task.ngay_bat_dau = datetime.fromisoformat(new_start_str)
        task.ngay_ket_thuc = datetime.fromisoformat(new_end_str)
        task.save()
        return JsonResponse({'status': 'success', 'message': 'Cập nhật thành công!'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
def task_handover_view(request):
    if request.method == 'POST':
        form = TaskHandoverForm(request.user, request.POST)
        if form.is_valid():
            task = form.cleaned_data['task']
            new_user = form.cleaned_data['new_user']
            task.nguoi_thuc_hien = new_user
            task.save()
            messages.success(request, f"Đã bàn giao nhiệm vụ '{task.ten_nhiem_vu}' cho {new_user.username} thành công!")
            return redirect('nhiemvu-list')
    else:
        form = TaskHandoverForm(user=request.user)
    context = {'form': form}
    return render(request, 'core/task_handover.html', context)

@login_required
def approve_hosocongviec_view(request, pk):
    hosocongviec = get_object_or_404(HoSoCongViec, pk=pk)
    if request.method == 'POST':
        if request.user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            hosocongviec.trang_thai = HoSoCongViec.TrangThai.DA_DUYET
            hosocongviec.save()
            messages.success(request, _(f'Hồ sơ công việc "{hosocongviec.ten_ho_so_cong_viec}" đã được phê duyệt.'))
        else:
            messages.error(request, _('Bạn không có quyền phê duyệt Hồ sơ công việc này.'))
    return redirect('hosocongviec-list')

@login_required
def approve_kehoach_view(request, pk):
    kehoach = get_object_or_404(KeHoach, pk=pk)
    if request.method == 'POST':
        if request.user.role == CustomUser.Role.LANH_DAO_PHONG:
            kehoach.trang_thai = KeHoach.TrangThai.DA_DUYET
            kehoach.save()
            messages.success(request, _(f'Kế hoạch "{kehoach.ten_ke_hoach}" đã được phê duyệt.'))
        else:
            messages.error(request, _('Bạn không có quyền phê duyệt Kế hoạch này.'))
    return redirect('kehoach-list')

@login_required
def approve_nhiemvu_view(request, pk):
    nhiemvu = get_object_or_404(NhiemVu, pk=pk)
    if request.method == 'POST':
        nhiemvu.trang_thai = NhiemVu.TrangThai.ASSIGNED
        nhiemvu.save()
        messages.success(request, _(f'Nhiệm vụ "{nhiemvu.ten_nhiem_vu}" đã được phê duyệt.'))
    return redirect('nhiemvu-list')

@login_required
def complete_and_rate_nhiemvu_view(request, pk):
    nhiemvu = get_object_or_404(NhiemVu, pk=pk)

    # Check if the current user is the main assignee
    if nhiemvu.id_nguoi_xu_ly_chinh != request.user:
        messages.error(request, "Bạn không có quyền thực hiện hành động này.")
        return redirect('nhiemvu-detail', pk=pk)

    if request.method == 'POST':
        form = NhiemVuCompletionForm(request.POST, request.FILES)
        if form.is_valid():
            # Handle file uploads
            for f in request.FILES.getlist('attachments'):
                TepDinhKem.objects.create(
                    file=f,
                    uploader=request.user,
                    nhiem_vu=nhiemvu
                )

            # Add comment if provided
            comment_text = form.cleaned_data.get('comment')
            if comment_text:
                BinhLuan.objects.create(
                    nhiem_vu=nhiemvu,
                    user=request.user,
                    noi_dung=comment_text
                )

            nhiemvu.trang_thai = NhiemVu.TrangThai.PENDING_COMPLETION_APPROVAL
            nhiemvu.save()
            messages.success(request, _(f'Nhiệm vụ "{nhiemvu.ten_nhiem_vu}" đã được gửi duyệt hoàn thành.'))

            # Add audit log entry
            LichSuCongViec.objects.create(
                nhiem_vu=nhiemvu,
                user=request.user,
                mo_ta="Gửi duyệt hoàn thành nhiệm vụ",
                details={
                    "trang_thai_cu": nhiemvu.trang_thai,
                    "trang_thai_moi": NhiemVu.TrangThai.PENDING_COMPLETION_APPROVAL,
                    "nguoi_gui_duyet": request.user.username,
                }
            )
            # Send notification to assigner
            message_to_assigner = f"Nhiệm vụ '{nhiemvu.ten_nhiem_vu}' đã được gửi duyệt hoàn thành bởi {request.user.username}."
            create_notification(nhiemvu.id_nguoi_giao_viec, message_to_assigner, related_task=nhiemvu)

            return redirect('nhiemvu-detail', pk=pk)
    else:
        form = NhiemVuCompletionForm()

    context = {
        'nhiemvu': nhiemvu,
        'form': form
    }
    return render(request, 'core/nhiemvu_completion_form.html', context)

@login_required
def approve_assignment_view(request, pk):
    nhiemvu = get_object_or_404(NhiemVu, pk=pk)

    # Check if the task is in PENDING_ASSIGNMENT_APPROVAL status
    if nhiemvu.trang_thai != NhiemVu.TrangThai.PENDING_ASSIGNMENT_APPROVAL:
        messages.error(request, "Nhiệm vụ này không ở trạng thái chờ phê duyệt giao.")
        return redirect('nhiemvu-detail', pk=pk)

    # Check if the current user has permission to approve
    # Lãnh đạo Cơ quan HOẶC Lãnh đạo Văn phòng
    if not (request.user.role == CustomUser.Role.LANH_DAO_CO_QUAN or \
            request.user.role == CustomUser.Role.LANH_DAO_VAN_PHONG):
        messages.error(request, "Bạn không có quyền phê duyệt nhiệm vụ này.")
        return redirect('nhiemvu-detail', pk=pk)

    if request.method == 'POST':
        nhiemvu.trang_thai = NhiemVu.TrangThai.ASSIGNED
        nhiemvu.id_nguoi_duyet_giao_viec = request.user  # Save the approver
        nhiemvu.save()
        messages.success(request, f"Đã phê duyệt giao nhiệm vụ '{nhiemvu.ten_nhiem_vu}' thành công!")
        
        # Add audit log entry for assignment approval
        LichSuCongViec.objects.create(
            nhiem_vu=nhiemvu,
            user=request.user,
            mo_ta="Phê duyệt giao nhiệm vụ",
            details={
                "trang_thai_cu": NhiemVu.TrangThai.PENDING_ASSIGNMENT_APPROVAL,
                "trang_thai_moi": NhiemVu.TrangThai.ASSIGNED,
                "nguoi_phe_duyet": request.user.username,
            }
        )
        # Send notification to assigner and assignee
        message_to_assigner = f"Nhiệm vụ '{nhiemvu.ten_nhiem_vu}' bạn đã giao đã được phê duyệt."
        create_notification(nhiemvu.id_nguoi_giao_viec, message_to_assigner, related_task=nhiemvu)

        message_to_assignee = f"Nhiệm vụ '{nhiemvu.ten_nhiem_vu}' đã được phê duyệt và giao cho bạn."
        create_notification(nhiemvu.id_nguoi_xu_ly_chinh, message_to_assignee, related_task=nhiemvu)

        return redirect('nhiemvu-detail', pk=pk)

    context = {'nhiemvu': nhiemvu}
    return render(request, 'core/nhiemvu_approve_assignment.html', context)

@login_required
def approve_completion_and_rate_nhiemvu_view(request, pk):
    nhiemvu = get_object_or_404(NhiemVu, pk=pk)

    # Check if the current user is the designated approver
    if request.user != nhiemvu.completion_approver:
        messages.error(request, "Bạn không có quyền phê duyệt và chấm điểm nhiệm vụ này.")
        return redirect('nhiemvu-detail', pk=pk)

    # Check if the task is in PENDING_COMPLETION_APPROVAL status
    if nhiemvu.trang_thai != NhiemVu.TrangThai.PENDING_COMPLETION_APPROVAL:
        messages.error(request, "Nhiệm vụ này không ở trạng thái chờ duyệt hoàn thành.")
        return redirect('nhiemvu-detail', pk=pk)

    if request.method == 'POST':
        rating = request.POST.get('danh_gia_sao')
        review = request.POST.get('loi_danh_gia')

        nhiemvu.trang_thai = NhiemVu.TrangThai.COMPLETED
        nhiemvu.danh_gia_sao = rating
        nhiemvu.loi_danh_gia = review
        nhiemvu.ngay_hoan_thanh = timezone.now() # Set completion date
        
        # Calculate actual business hours
        if nhiemvu.ngay_bat_dau and nhiemvu.ngay_hoan_thanh:
            print(f"DEBUG: ngay_bat_dau: {nhiemvu.ngay_bat_dau}, ngay_hoan_thanh: {nhiemvu.ngay_hoan_thanh}")
            calculated_time = calculate_business_hours(nhiemvu.ngay_bat_dau, nhiemvu.ngay_hoan_thanh)
            nhiemvu.thoi_gian_thuc_te = calculated_time
            print(f"DEBUG: Calculated thoi_gian_thuc_te: {calculated_time}")
        else:
            print(f"DEBUG: Cannot calculate thoi_gian_thuc_te. ngay_bat_dau: {nhiemvu.ngay_bat_dau}, ngay_hoan_thanh: {nhiemvu.ngay_hoan_thanh}")

        nhiemvu.save()

        messages.success(request, f"Đã phê duyệt hoàn thành và chấm điểm nhiệm vụ '{nhiemvu.ten_nhiem_vu}' thành công!")

        # Add audit log entry
        LichSuCongViec.objects.create(
            nhiem_vu=nhiemvu,
            user=request.user,
            mo_ta="Phê duyệt hoàn thành và chấm điểm nhiệm vụ",
            details={
                "trang_thai_cu": NhiemVu.TrangThai.PENDING_COMPLETION_APPROVAL,
                "trang_thai_moi": NhiemVu.TrangThai.COMPLETED,
                "nguoi_phe_duyet": request.user.username,
                "danh_gia_sao": rating,
                "loi_danh_gia": review,
            }
        )
        # Send notification to assignee
        message_to_assignee = f"Nhiệm vụ '{nhiemvu.ten_nhiem_vu}' của bạn đã được phê duyệt hoàn thành và chấm điểm."
        create_notification(nhiemvu.id_nguoi_xu_ly_chinh, message_to_assignee, related_task=nhiemvu)

        return redirect('nhiemvu-detail', pk=pk)

    # For GET request, render a template to allow rating
    context = {
        'nhiemvu': nhiemvu,
        'default_rating': 5 # As per v6.md, default rating is 5 stars
    }
    return render(request, 'core/nhiemvu_approve_completion.html', context)

@login_required
def request_extension_view(request, pk):
    nhiemvu = get_object_or_404(NhiemVu, pk=pk)

    # Check if the current user is the main assignee of the task
    if nhiemvu.id_nguoi_xu_ly_chinh != request.user:
        messages.error(request, "Bạn không có quyền yêu cầu gia hạn cho nhiệm vụ này.")
        return redirect('nhiemvu-detail', pk=pk)

    # Check if the task is in ASSIGNED status (or other relevant "in progress" status)
    if nhiemvu.trang_thai != NhiemVu.TrangThai.ASSIGNED:
        messages.error(request, "Chỉ có thể yêu cầu gia hạn cho nhiệm vụ đang được giao.")
        return redirect('nhiemvu-detail', pk=pk)

    if request.method == 'POST':
        form = ExtensionRequestForm(request.POST)
        if form.is_valid():
            nhiemvu.ngay_ket_thuc_de_xuat = form.cleaned_data['ngay_ket_thuc_de_xuat']
            nhiemvu.ly_do_gia_han = form.cleaned_data['ly_do_gia_han']
            nhiemvu.trang_thai_gia_han = NhiemVu.TrangThaiGiaHan.CHUA_DUYET
            nhiemvu.save()

            messages.success(request, "Yêu cầu gia hạn đã được gửi thành công.")

            # Add audit log entry
            LichSuCongViec.objects.create(
                nhiem_vu=nhiemvu,
                user=request.user,
                mo_ta="Yêu cầu gia hạn nhiệm vụ",
                details={
                    "ngay_ket_thuc_cu": nhiemvu.ngay_ket_thuc.isoformat(),
                    "ngay_ket_thuc_de_xuat": nhiemvu.ngay_ket_thuc_de_xuat.isoformat(),
                    "ly_do": nhiemvu.ly_do_gia_han,
                }
            )

            # Send notification to the assigner
            message_to_assigner = f"Nhiệm vụ '{nhiemvu.ten_nhiem_vu}' có yêu cầu gia hạn từ {request.user.username}."
            create_notification(nhiemvu.id_nguoi_giao_viec, message_to_assigner, related_task=nhiemvu)

            return redirect('nhiemvu-detail', pk=pk)
    else:
        form = ExtensionRequestForm()

    context = {
        'nhiemvu': nhiemvu,
        'form': form
    }
    return render(request, 'core/nhiemvu_request_extension.html', context)

# --- Class-Based Views (CRUD) ---
# CoQuan Views
class CoQuanListView(LoginRequiredMixin, ListView):
    model = CoQuan
    template_name = 'core/coquan_list.html'

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            return CoQuan.objects.all()
        elif user.co_quan:
            return CoQuan.objects.filter(pk=user.co_quan.pk)
        else:
            return CoQuan.objects.none()
class CoQuanDetailView(LoginRequiredMixin, DetailView):
    model = CoQuan
    template_name = 'core/coquan_detail.html'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        user = self.request.user

        if user.is_superuser or user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            pass
        elif user.co_quan and obj.pk == user.co_quan.pk:
            pass
        else:
            raise Http404("Cơ quan không tồn tại hoặc bạn không có quyền truy cập.")
        return obj
class CoQuanCreateView(LoginRequiredMixin, CreateView):
    model = CoQuan
    form_class = CoQuanForm
    template_name = 'core/coquan_form.html'
    success_url = reverse_lazy('coquan-list')
class CoQuanUpdateView(LoginRequiredMixin, UpdateView):
    model = CoQuan
    form_class = CoQuanForm
    template_name = 'core/coquan_form.html'
    success_url = reverse_lazy('coquan-list')
class CoQuanDeleteView(LoginRequiredMixin, DeleteView):
    model = CoQuan
    template_name = 'core/coquan_confirm_delete.html'
    success_url = reverse_lazy('coquan-list')

# PhongBan Views
class PhongBanListView(LoginRequiredMixin, ListView):
    model = PhongBan
    template_name = 'core/phongban_list.html'

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            return PhongBan.objects.all()
        elif user.role == CustomUser.Role.LANH_DAO_VAN_PHONG:
            if user.phong_ban and user.phong_ban.co_quan:
                return PhongBan.objects.filter(co_quan=user.phong_ban.co_quan)
            else:
                return PhongBan.objects.none()
        elif user.role == CustomUser.Role.LANH_DAO_PHONG or user.role == CustomUser.Role.CHUYEN_VIEN_VAN_PHONG or user.role == CustomUser.Role.CHUYEN_VIEN_PHONG:
            if user.phong_ban:
                return PhongBan.objects.filter(pk=user.phong_ban.pk)
            else:
                return PhongBan.objects.none()
        else:
            return PhongBan.objects.none()
class PhongBanDetailView(LoginRequiredMixin, DetailView):
    model = PhongBan
    template_name = 'core/phongban_detail.html'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        user = self.request.user

        if user.is_superuser or user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            pass
        elif user.role == CustomUser.Role.LANH_DAO_VAN_PHONG:
            if user.phong_ban and user.phong_ban.co_quan and obj.co_quan == user.phong_ban.co_quan:
                pass
            else:
                raise Http404("Phòng ban không tồn tại hoặc bạn không có quyền truy cập.")
        elif user.role == CustomUser.Role.LANH_DAO_PHONG or user.role == CustomUser.Role.CHUYEN_VIEN_VAN_PHONG or user.role == CustomUser.Role.CHUYEN_VIEN_PHONG:
            if user.phong_ban and obj.pk == user.phong_ban.pk:
                pass
            else:
                raise Http404("Phòng ban không tồn tại hoặc bạn không có quyền truy cập.")
        else:
            raise Http404("Phòng ban không tồn tại hoặc bạn không có quyền truy cập.")
        return obj
class PhongBanCreateView(LoginRequiredMixin, CreateView):
    model = PhongBan
    form_class = PhongBanForm
    template_name = 'core/phongban_form.html'
    success_url = reverse_lazy('phongban-list')
class PhongBanUpdateView(LoginRequiredMixin, UpdateView):
    model = PhongBan
    form_class = PhongBanForm
    template_name = 'core/phongban_form.html'
    success_url = reverse_lazy('phongban-list')
class PhongBanDeleteView(LoginRequiredMixin, DeleteView):
    model = PhongBan
    template_name = 'core/phongban_confirm_delete.html'
    success_url = reverse_lazy('phongban-list')

# HoSoCongViec Views
class HoSoCongViecListView(LoginRequiredMixin, ListView):
    model = HoSoCongViec
    template_name = 'core/hosocongviec_list.html'

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            return HoSoCongViec.objects.all()
        elif user.role == CustomUser.Role.LANH_DAO_VAN_PHONG:
            if user.phong_ban and user.phong_ban.co_quan:
                return HoSoCongViec.objects.filter(id_don_vi_chu_tri__co_quan=user.phong_ban.co_quan)
            else:
                return HoSoCongViec.objects.none()
        elif user.role == CustomUser.Role.LANH_DAO_PHONG:
            if user.phong_ban:
                return HoSoCongViec.objects.filter(id_don_vi_chu_tri=user.phong_ban)
            else:
                return HoSoCongViec.objects.none()
        elif user.role == CustomUser.Role.CHUYEN_VIEN_VAN_PHONG or user.role == CustomUser.Role.CHUYEN_VIEN_PHONG:
            return HoSoCongViec.objects.filter(id_nguoi_quan_ly=user) | HoSoCongViec.objects.filter(id_don_vi_chu_tri=user.phong_ban)
        else:
            return HoSoCongViec.objects.none()
class HoSoCongViecDetailView(LoginRequiredMixin, DetailView):
    model = HoSoCongViec
    template_name = 'core/hosocongviec_detail.html'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        user = self.request.user

        if user.is_superuser or user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            pass
        elif user.role == CustomUser.Role.LANH_DAO_VAN_PHONG:
            if user.phong_ban and user.phong_ban.co_quan and obj.id_don_vi_chu_tri and obj.id_don_vi_chu_tri.co_quan == user.phong_ban.co_quan:
                pass
            else:
                raise Http404("Hồ sơ công việc không tồn tại hoặc bạn không có quyền truy cập.")
        elif user.role == CustomUser.Role.LANH_DAO_PHONG:
            if user.phong_ban and obj.id_don_vi_chu_tri == user.phong_ban:
                pass
            else:
                raise Http404("Hồ sơ công việc không tồn tại hoặc bạn không có quyền truy cập.")
        elif user.role == CustomUser.Role.CHUYEN_VIEN_VAN_PHONG or user.role == CustomUser.Role.CHUYEN_VIEN_PHONG:
            if obj.id_nguoi_quan_ly == user or (user.phong_ban and obj.id_don_vi_chu_tri == user.phong_ban):
                pass
            else:
                raise Http404("Hồ sơ công việc không tồn tại hoặc bạn không có quyền truy cập.")
        else:
            raise Http404("Hồ sơ công việc không tồn tại hoặc bạn không có quyền truy cập.")
        return obj
class HoSoCongViecCreateView(LoginRequiredMixin, CreateView):
    model = HoSoCongViec
    form_class = HoSoCongViecForm
    template_name = 'core/hosocongviec_form.html'
    success_url = reverse_lazy('hosocongviec-list')

    def form_valid(self, form):
        response = super().form_valid(form)

        # Handle file uploads
        for f in self.request.FILES.getlist('attachments'):
            TepDinhKem.objects.create(
                file=f,
                uploader=self.request.user,
                ho_so_cong_viec=form.instance
            )
        messages.success(self.request, "Hồ sơ công việc và tệp đính kèm đã được tạo thành công!")

        return response
class HoSoCongViecUpdateView(LoginRequiredMixin, UpdateView):
    model = HoSoCongViec
    form_class = HoSoCongViecForm
    template_name = 'core/hosocongviec_form.html'
    success_url = reverse_lazy('hosocongviec-list')

    def form_valid(self, form):
        response = super().form_valid(form)

        # Handle file uploads
        for f in self.request.FILES.getlist('attachments'):
            TepDinhKem.objects.create(
                file=f,
                uploader=self.request.user,
                ho_so_cong_viec=form.instance
            )
        messages.success(self.request, "Hồ sơ công việc và tệp đính kèm đã được cập nhật thành công!")

        return response
class HoSoCongViecDeleteView(LoginRequiredMixin, DeleteView):
    model = HoSoCongViec
    template_name = 'core/hosocongviec_confirm_delete.html'
    success_url = reverse_lazy('hosocongviec-list')

# KeHoach Views
class KeHoachListView(LoginRequiredMixin, ListView):
    model = KeHoach
    template_name = 'core/kehoach_list.html'

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            return KeHoach.objects.all()
        elif user.role == CustomUser.Role.LANH_DAO_VAN_PHONG:
            if user.phong_ban and user.phong_ban.co_quan:
                return KeHoach.objects.filter(id_don_vi_thuc_hien__co_quan=user.phong_ban.co_quan)
            else:
                return KeHoach.objects.none()
        elif user.role == CustomUser.Role.LANH_DAO_PHONG:
            if user.phong_ban:
                return KeHoach.objects.filter(id_don_vi_thuc_hien=user.phong_ban)
            else:
                return KeHoach.objects.none()
        elif user.role == CustomUser.Role.CHUYEN_VIEN_VAN_PHONG or user.role == CustomUser.Role.CHUYEN_VIEN_PHONG:
            return KeHoach.objects.filter(id_nguoi_phu_trach=user) | KeHoach.objects.filter(id_don_vi_thuc_hien=user.phong_ban)
        else:
            return KeHoach.objects.none()
class KeHoachDetailView(LoginRequiredMixin, DetailView):
    model = KeHoach
    template_name = 'core/kehoach_detail.html'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        user = self.request.user

        if user.is_superuser or user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            pass
        elif user.role == CustomUser.Role.LANH_DAO_VAN_PHONG:
            if user.phong_ban and user.phong_ban.co_quan and obj.id_don_vi_thuc_hien and obj.id_don_vi_thuc_hien.co_quan == user.phong_ban.co_quan:
                pass
            else:
                raise Http404("Kế hoạch không tồn tại hoặc bạn không có quyền truy cập.")
        elif user.role == CustomUser.Role.LANH_DAO_PHONG:
            if user.phong_ban and obj.id_don_vi_thuc_hien == user.phong_ban:
                pass
            else:
                raise Http404("Kế hoạch không tồn tại hoặc bạn không có quyền truy cập.")
        elif user.role == CustomUser.Role.CHUYEN_VIEN_VAN_PHONG or user.role == CustomUser.Role.CHUYEN_VIEN_PHONG:
            if obj.id_nguoi_phu_trach == user or (user.phong_ban and obj.id_don_vi_thuc_hien == user.phong_ban):
                pass
            else:
                raise Http404("Kế hoạch không tồn tại hoặc bạn không có quyền truy cập.")
        else:
            raise Http404("Kế hoạch không tồn tại hoặc bạn không có quyền truy cập.")
        return obj
class KeHoachCreateView(LoginRequiredMixin, CreateView):
    model = KeHoach
    form_class = KeHoachForm
    template_name = 'core/kehoach_form.html'
    success_url = reverse_lazy('kehoach-list')

    def form_valid(self, form):
        response = super().form_valid(form)

        # Handle file uploads
        for f in self.request.FILES.getlist('attachments'):
            TepDinhKem.objects.create(
                file=f,
                uploader=self.request.user,
                ke_hoach=form.instance
            )
        messages.success(self.request, "Kế hoạch và tệp đính kèm đã được tạo thành công!")

        return response
class KeHoachUpdateView(LoginRequiredMixin, UpdateView):
    model = KeHoach
    form_class = KeHoachForm
    template_name = 'core/kehoach_form.html'
    success_url = reverse_lazy('kehoach-list')

    def form_valid(self, form):
        response = super().form_valid(form)

        # Handle file uploads
        for f in self.request.FILES.getlist('attachments'):
            TepDinhKem.objects.create(
                file=f,
                uploader=self.request.user,
                ke_hoach=form.instance
            )
        messages.success(self.request, "Kế hoạch và tệp đính kèm đã được cập nhật thành công!")

        return response
class KeHoachDeleteView(LoginRequiredMixin, DeleteView):
    model = KeHoach
    template_name = 'core/kehoach_confirm_delete.html'
    success_url = reverse_lazy('kehoach-list')

# MocThoiGian Views
class MocThoiGianListView(LoginRequiredMixin, ListView):
    model = MocThoiGian
    template_name = 'core/mocthoigian_list.html'

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            return MocThoiGian.objects.all()
        elif user.role == CustomUser.Role.LANH_DAO_VAN_PHONG:
            if user.phong_ban and user.phong_ban.co_quan:
                return MocThoiGian.objects.filter(id_ke_hoach__id_don_vi_thuc_hien__co_quan=user.phong_ban.co_quan)
            else:
                return MocThoiGian.objects.none()
        elif user.role == CustomUser.Role.LANH_DAO_PHONG:
            if user.phong_ban:
                return MocThoiGian.objects.filter(id_ke_hoach__id_don_vi_thuc_hien=user.phong_ban)
            else:
                return MocThoiGian.objects.none()
        elif user.role == CustomUser.Role.CHUYEN_VIEN_VAN_PHONG or user.role == CustomUser.Role.CHUYEN_VIEN_PHONG:
            # Chuyen vien sees milestones related to plans they are responsible for or their department is involved in
            return MocThoiGian.objects.filter(id_ke_hoach__id_nguoi_phu_trach=user) | MocThoiGian.objects.filter(id_ke_hoach__id_don_vi_thuc_hien=user.phong_ban)
        else:
            return MocThoiGian.objects.filter(id_ke_hoach__nhiem_vu__id_nguoi_thuc_hien=user) # Fallback for other roles, assuming they only see tasks assigned to them
class MocThoiGianDetailView(LoginRequiredMixin, DetailView):
    model = MocThoiGian
    template_name = 'core/mocthoigian_detail.html'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        user = self.request.user

        if user.is_superuser or user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            pass
        elif user.role == CustomUser.Role.LANH_DAO_VAN_PHONG:
            if user.phong_ban and user.phong_ban.co_quan and obj.id_ke_hoach.id_don_vi_thuc_hien and obj.id_ke_hoach.id_don_vi_thuc_hien.co_quan == user.phong_ban.co_quan:
                pass
            else:
                raise Http404("Mốc thời gian không tồn tại hoặc bạn không có quyền truy cập.")
        elif user.role == CustomUser.Role.LANH_DAO_PHONG:
            if user.phong_ban and obj.id_ke_hoach.id_don_vi_thuc_hien == user.phong_ban:
                pass
            else:
                raise Http404("Mốc thời gian không tồn tại hoặc bạn không có quyền truy cập.")
        elif user.role == CustomUser.Role.CHUYEN_VIEN_VAN_PHONG or user.role == CustomUser.Role.CHUYEN_VIEN_PHONG:
            if obj.id_ke_hoach.id_nguoi_phu_trach == user or (user.phong_ban and obj.id_ke_hoach.id_don_vi_thuc_hien == user.phong_ban):
                pass
            else:
                raise Http404("Mốc thời gian không tồn tại hoặc bạn không có quyền truy cập.")
        else:
            raise Http404("Mốc thời gian không tồn tại hoặc bạn không có quyền truy cập.")
        return obj
class MocThoiGianCreateView(LoginRequiredMixin, CreateView):
    model = MocThoiGian
    form_class = MocThoiGianForm
    template_name = 'core/mocthoigian_form.html'
    success_url = reverse_lazy('mocthoigian-list')
class MocThoiGianUpdateView(LoginRequiredMixin, UpdateView):
    model = MocThoiGian
    form_class = MocThoiGianForm
    template_name = 'core/mocthoigian_form.html'
    success_url = reverse_lazy('mocthoigian-list')
class MocThoiGianDeleteView(LoginRequiredMixin, DeleteView):
    model = MocThoiGian
    template_name = 'core/mocthoigian_confirm_delete.html'
    success_url = reverse_lazy('mocthoigian-list')

# NhiemVu Views
class NhiemVuListView(LoginRequiredMixin, ListView):
    model = NhiemVu
    template_name = 'core/nhiemvu_list.html'

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            # Lãnh đạo Cơ quan (Global View): Thấy tất cả nhiệm vụ.
            return NhiemVu.objects.all()
        elif user.role == CustomUser.Role.LANH_DAO_VAN_PHONG:
            # Lãnh đạo Văn phòng (Office-scoped View): Thấy các nhiệm vụ liên quan đến cơ quan của mình.
            # Assuming Lanh dao Van phong is associated with a CoQuan via their PhongBan
            if user.phong_ban and user.phong_ban.co_quan:
                return NhiemVu.objects.filter(
                    id_phong_ban_lien_quan__co_quan=user.phong_ban.co_quan
                )
            else:
                return NhiemVu.objects.none()
        elif user.role == CustomUser.Role.LANH_DAO_PHONG:
            # Lãnh đạo Phòng (Department-scoped View): Chỉ thấy nhiệm vụ trong phòng ban của mình.
            if user.phong_ban:
                return NhiemVu.objects.filter(id_phong_ban_lien_quan=user.phong_ban)
            else:
                return NhiemVu.objects.none()
        elif user.role == CustomUser.Role.CHUYEN_VIEN_VAN_PHONG or user.role == CustomUser.Role.CHUYEN_VIEN_PHONG:
            # Chuyên viên (Personal View): Chỉ thấy nhiệm vụ của bản thân (người xử lý chính hoặc người giao việc hoặc người đồng xử lý).
            return NhiemVu.objects.filter(Q(id_nguoi_xu_ly_chinh=user) | Q(id_nguoi_giao_viec=user) | Q(nguoi_dong_xu_ly=user)).distinct()
        else:
            return NhiemVu.objects.none()
class NhiemVuDetailView(LoginRequiredMixin, DetailView):
    model = NhiemVu
    template_name = 'core/nhiemvu_detail.html'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        user = self.request.user

        if user.is_superuser or user.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            # Lãnh đạo Cơ quan (Global View): Thấy tất cả nhiệm vụ.
            pass # No additional filtering needed
        elif user.role == CustomUser.Role.LANH_DAO_VAN_PHONG:
            # Lãnh đạo Văn phòng (Office-scoped View): Thấy các nhiệm vụ liên quan đến cơ quan của mình.
            if user.phong_ban and user.phong_ban.co_quan and obj.id_phong_ban_lien_quan and obj.id_phong_ban_lien_quan.co_quan == user.phong_ban.co_quan:
                pass
            else:
                raise Http404("Nhiệm vụ không tồn tại hoặc bạn không có quyền truy cập.")
        elif user.role == CustomUser.Role.LANH_DAO_PHONG:
            # Lãnh đạo Phòng (Department-scoped View): Chỉ thấy nhiệm vụ trong phòng ban của mình.
            if user.phong_ban and obj.id_phong_ban_lien_quan == user.phong_ban:
                pass
            else:
                raise Http404("Nhiệm vụ không tồn tại hoặc bạn không có quyền truy cập.")
        elif user.role == CustomUser.Role.CHUYEN_VIEN_VAN_PHONG or user.role == CustomUser.Role.CHUYEN_VIEN_PHONG:
            # Chuyên viên (Personal View): Chỉ thấy nhiệm vụ của bản thân (người xử lý chính hoặc người giao việc hoặc người đồng xử lý).
            if obj.id_nguoi_xu_ly_chinh == user or obj.id_nguoi_giao_viec == user or user in obj.nguoi_dong_xu_ly.all():
                pass
            else:
                raise Http404("Nhiệm vụ không tồn tại hoặc bạn không có quyền truy cập.")
        else:
            raise Http404("Nhiệm vụ không tồn tại hoặc bạn không có quyền truy cập.")
        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['binh_luan'] = self.object.binh_luan.all() # Get all comments for the task
        context['form'] = BinhLuanForm() # Pass an empty form for comments
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object() # Get the task object
        form = BinhLuanForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.nhiem_vu = self.object
            comment.user = request.user
            comment.save()
            messages.success(request, "Bình luận của bạn đã được thêm.")
            return redirect('nhiemvu-detail', pk=self.object.pk)
        else:
            # If form is not valid, re-render the page with errors
            context = self.get_context_data()
            context['form'] = form # Pass the form with errors
            return self.render_to_response(context)
class NhiemVuCreateView(LoginRequiredMixin, CreateView):
    model = NhiemVu
    form_class = NhiemVuForm
    template_name = 'core/nhiemvu_form.html'
    success_url = reverse_lazy('nhiemvu-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Pass the current user to the form
        return kwargs

    def form_valid(self, form):
        # Set creator and assigner
        form.instance.id_nguoi_tao = self.request.user
        form.instance.id_nguoi_giao_viec = self.request.user

        # Ensure ngay_bat_dau is set and timezone-aware
        if not form.instance.ngay_bat_dau:
            form.instance.ngay_bat_dau = timezone.now()
        else:
            # Make it timezone-aware if it's naive (from form input)
            if form.instance.ngay_bat_dau.tzinfo is None:
                form.instance.ngay_bat_dau = timezone.make_aware(form.instance.ngay_bat_dau)

        # Make ngay_ket_thuc timezone-aware if it's naive
        if form.instance.ngay_ket_thuc and form.instance.ngay_ket_thuc.tzinfo is None:
            form.instance.ngay_ket_thuc = timezone.make_aware(form.instance.ngay_ket_thuc)

        # Determine initial status based on assignment rules
        main_assignee = form.cleaned_data['id_nguoi_xu_ly_chinh']
        co_assignees = form.cleaned_data['nguoi_dong_xu_ly']

        # Check if main_assignee is a Super Admin
        if main_assignee.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            messages.error(self.request, "Không thể giao việc cho vai trò Lãnh đạo Cơ quan.")
            return self.form_invalid(form)

        if self.request.user.role == CustomUser.Role.CHUYEN_VIEN_VAN_PHONG and \
           main_assignee.role == CustomUser.Role.LANH_DAO_PHONG:
            form.instance.trang_thai = NhiemVu.TrangThai.PENDING_ASSIGNMENT_APPROVAL
        else:
            form.instance.trang_thai = NhiemVu.TrangThai.ASSIGNED

        # Set id_phong_ban_lien_quan based on main_assignee's department
        if main_assignee.phong_ban:
            form.instance.id_phong_ban_lien_quan = main_assignee.phong_ban

        response = super().form_valid(form)

        # Save many-to-many relationship for co_assignees
        form.instance.nguoi_dong_xu_ly.set(co_assignees)

        # Handle file uploads
        for f in self.request.FILES.getlist('attachments'):
            TepDinhKem.objects.create(
                file=f,
                uploader=self.request.user,
                nhiem_vu=form.instance
            )
        messages.success(self.request, "Nhiệm vụ và tệp đính kèm đã được tạo thành công!")

        # Add audit log entry for task creation
        LichSuCongViec.objects.create(
            nhiem_vu=form.instance,
            user=self.request.user,
            mo_ta="Tạo nhiệm vụ mới",
            details={
                "trang_thai_ban_dau": form.instance.trang_thai,
                "nguoi_tao": self.request.user.username,
                "nguoi_giao_viec": form.instance.id_nguoi_giao_viec.username,
                "nguoi_xu_ly_chinh": form.instance.id_nguoi_xu_ly_chinh.username,
                "nguoi_dong_xu_ly": [user.username for user in form.instance.nguoi_dong_xu_ly.all()],
            }
        )

        # Send notifications for PENDING_ASSIGNMENT_APPROVAL
        if form.instance.trang_thai == NhiemVu.TrangThai.PENDING_ASSIGNMENT_APPROVAL:
            message = f"Nhiệm vụ '{form.instance.ten_nhiem_vu}' cần được phê duyệt giao."
            # Notify Lãnh đạo Cơ quan
            for user in CustomUser.objects.filter(role=CustomUser.Role.LANH_DAO_CO_QUAN):
                create_notification(user, message, related_task=form.instance)
            # Notify Lãnh đạo Văn phòng
            for user in CustomUser.objects.filter(role=CustomUser.Role.LANH_DAO_VAN_PHONG):
                create_notification(user, message, related_task=form.instance)
        else: # Task is assigned directly
            # Send Zalo message to main assignee
            if form.instance.id_nguoi_xu_ly_chinh and form.instance.id_nguoi_xu_ly_chinh.zalo_user_id:
                zalo_message = f"Bạn có nhiệm vụ mới: {form.instance.ten_nhiem_vu}. Mô tả: {form.instance.mo_ta[:100]}..."
                send_zalo_message(form.instance.id_nguoi_xu_ly_chinh.zalo_user_id, zalo_message)
            
            # Send Zalo message to co-assignees
            for co_assignee in form.instance.nguoi_dong_xu_ly.all():
                if co_assignee.zalo_user_id:
                    zalo_message = f"Bạn được thêm vào nhiệm vụ: {form.instance.ten_nhiem_vu} với vai trò đồng xử lý. Mô tả: {form.instance.mo_ta[:100]}..."
                    send_zalo_message(co_assignee.zalo_user_id, zalo_message)

        return response
class NhiemVuUpdateView(LoginRequiredMixin, UpdateView):
    model = NhiemVu
    form_class = NhiemVuForm
    template_name = 'core/nhiemvu_form.html'
    success_url = reverse_lazy('nhiemvu-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Pass the current user to the form
        return kwargs

    def form_valid(self, form):
        # Logic to handle status changes on update, similar to CreateView
        original_instance = self.get_object()
        new_instance = form.save(commit=False)

        # Ensure ngay_bat_dau is set and timezone-aware
        if not new_instance.ngay_bat_dau:
            new_instance.ngay_bat_dau = timezone.now()
        else:
            # Make it timezone-aware if it's naive (from form input)
            if new_instance.ngay_bat_dau.tzinfo is None:
                new_instance.ngay_bat_dau = timezone.make_aware(new_instance.ngay_bat_dau)

        # Make ngay_ket_thuc timezone-aware if it's naive
        if new_instance.ngay_ket_thuc and new_instance.ngay_ket_thuc.tzinfo is None:
            new_instance.ngay_ket_thuc = timezone.make_aware(new_instance.ngay_ket_thuc)

        # Determine initial status based on assignment rules
        main_assignee = new_instance.id_nguoi_xu_ly_chinh
        co_assignees = form.cleaned_data['nguoi_dong_xu_ly']

        # Check if main_assignee is a Super Admin
        if main_assignee.role == CustomUser.Role.LANH_DAO_CO_QUAN:
            messages.error(self.request, "Không thể giao việc cho vai trò Lãnh đạo Cơ quan.")
            return self.form_invalid(form)

        # If the assignee or special status changes, re-evaluate the task status
        if self.request.user.role == CustomUser.Role.CHUYEN_VIEN_VAN_PHONG and \
           main_assignee.role == CustomUser.Role.LANH_DAO_PHONG:
            new_instance.trang_thai = NhiemVu.TrangThai.PENDING_ASSIGNMENT_APPROVAL
            # Reset assignment approver if the flow changes back to pending
            new_instance.id_nguoi_duyet_giao_viec = None
        else:
            # If it's not the special flow, it becomes ASSIGNED directly
            new_instance.trang_thai = NhiemVu.TrangThai.ASSIGNED
            # A non-special-flow task has no separate assignment approver
            new_instance.id_nguoi_duyet_giao_viec = None

        # Set id_phong_ban_lien_quan based on main_assignee's department
        if main_assignee.phong_ban:
            new_instance.id_phong_ban_lien_quan = main_assignee.phong_ban

        response = super().form_valid(form)

        # Save many-to-many relationship for co_assignees
        form.instance.nguoi_dong_xu_ly.set(co_assignees)

        # Handle file uploads
        for f in self.request.FILES.getlist('attachments'):
            TepDinhKem.objects.create(
                file=f,
                uploader=self.request.user,
                nhiem_vu=self.object
            )
        messages.success(self.request, "Nhiệm vụ và tệp đính kèm đã được cập nhật thành công!")

        # Send Zalo message to main assignee if changed or task status changed to ASSIGNED
        if new_instance.trang_thai == NhiemVu.TrangThai.ASSIGNED and \
           (original_instance.id_nguoi_xu_ly_chinh != new_instance.id_nguoi_xu_ly_chinh or \
            original_instance.trang_thai != NhiemVu.TrangThai.ASSIGNED):
            if new_instance.id_nguoi_xu_ly_chinh and new_instance.id_nguoi_xu_ly_chinh.zalo_user_id:
                zalo_message = f"Nhiệm vụ của bạn đã được cập nhật: {new_instance.ten_nhiem_vu}. Mô tả: {new_instance.mo_ta[:100]}..."
                send_zalo_message(new_instance.id_nguoi_xu_ly_chinh.zalo_user_id, zalo_message)
        
        # Send Zalo message to co-assignees if changed
        current_co_assignees = set(original_instance.nguoi_dong_xu_ly.all())
        new_co_assignees = set(new_instance.nguoi_dong_xu_ly.all())

        added_co_assignees = new_co_assignees - current_co_assignees
        for co_assignee in added_co_assignees:
            if co_assignee.zalo_user_id:
                zalo_message = f"Bạn được thêm vào nhiệm vụ: {new_instance.ten_nhiem_vu} với vai trò đồng xử lý. Mô tả: {new_instance.mo_ta[:100]}..."
                send_zalo_message(co_assignee.zalo_user_id, zalo_message)

        return response

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object() # Get the task object
        # Check if the current user is the main assignee or a co-assignee
        if self.object.id_nguoi_xu_ly_chinh == request.user or request.user in self.object.nguoi_dong_xu_ly.all():
            messages.error(request, "Bạn không có quyền sửa đổi nhiệm vụ này.")
            return redirect('nhiemvu-detail', pk=self.object.pk) # Redirect to detail view
        return super().dispatch(request, *args, **kwargs)
class NhiemVuDeleteView(LoginRequiredMixin, DeleteView):
    model = NhiemVu
    template_name = 'core/nhiemvu_confirm_delete.html'
    success_url = reverse_lazy('nhiemvu-list')

# TepDinhKem Views
class TepDinhKemListView(LoginRequiredMixin, ListView):
    model = TepDinhKem
    template_name = 'core/tepdinhkem_list.html'
class TepDinhKemDetailView(LoginRequiredMixin, DetailView):
    model = TepDinhKem
    template_name = 'core/tepdinhkem_detail.html'
class TepDinhKemCreateView(LoginRequiredMixin, CreateView):
    model = TepDinhKem
    form_class = TepDinhKemForm
    template_name = 'core/tepdinhkem_form.html'
    success_url = reverse_lazy('tepdinhkem-list')
class TepDinhKemUpdateView(LoginRequiredMixin, UpdateView):
    model = TepDinhKem
    form_class = TepDinhKemForm
    template_name = 'core/tepdinhkem_form.html'
    success_url = reverse_lazy('tepdinhkem-list')
class TepDinhKemDeleteView(LoginRequiredMixin, DeleteView):
    model = TepDinhKem
    template_name = 'core/tepdinhkem_confirm_delete.html'
    success_url = reverse_lazy('tepdinhkem-list')

# LichSuCongViec Views
class LichSuCongViecListView(LoginRequiredMixin, ListView):
    model = LichSuCongViec
    template_name = 'core/lichsucongviec_list.html'
class LichSuCongViecDetailView(LoginRequiredMixin, DetailView):
    model = LichSuCongViec
    template_name = 'core/lichsucongviec_detail.html'
class LichSuCongViecCreateView(LoginRequiredMixin, CreateView):
    model = LichSuCongViec
    form_class = LichSuCongViecForm
    template_name = 'core/lichsucongviec_form.html'
    success_url = reverse_lazy('lichsucongviec-list')
class LichSuCongViecUpdateView(LoginRequiredMixin, UpdateView):
    model = LichSuCongViec
    form_class = LichSuCongViecForm
    template_name = 'core/lichsucongviec_form.html'
    success_url = reverse_lazy('lichsucongviec-list')
class LichSuCongViecDeleteView(LoginRequiredMixin, DeleteView):
    model = LichSuCongViec
    template_name = 'core/lichsucongviec_confirm_delete.html'
    success_url = reverse_lazy('lichsucongviec-list')

# CustomReport Views
class CustomReportListView(LoginRequiredMixin, ListView):
    model = CustomReport
    template_name = 'core/customreport_list.html'
class CustomReportDetailView(LoginRequiredMixin, DetailView):
    model = CustomReport
    template_name = 'core/customreport_detail.html'
class CustomReportCreateView(LoginRequiredMixin, CreateView):
    model = CustomReport
    form_class = CustomReportForm
    template_name = 'core/customreport_form.html'
    success_url = reverse_lazy('customreport-list')
class CustomReportUpdateView(LoginRequiredMixin, UpdateView):
    model = CustomReport
    form_class = CustomReportForm
    template_name = 'core/customreport_form.html'
    success_url = reverse_lazy('customreport-list')
class CustomReportDeleteView(LoginRequiredMixin, DeleteView):
    model = CustomReport
    template_name = 'core/customreport_confirm_delete.html'
    success_url = reverse_lazy('customreport-list')